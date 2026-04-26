from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook


RAW_RESPONSE_TO_MAPPED = {
    "Категорически не хочу": "Не хочу",
    "Скорее нет": "Не хочу",
    "Не уверен": "Не уверен",
    "Скорее да": "Хочу",
    "Однозначно хочу": "Хочу",
    "Я не знаю эту компанию": "Я не знаю эту компанию",
}

METRIC_KIND_MAP = {
    "Сортировка по желанию работать в компании": "want",
    "Сортировка по не желанию работать в компании": "dont_want",
    "Сортировка по знанию бренда": "brand_awareness",
    "Сортировка по не уверенности респондентов работать в компании": "uncertainty",
}

METRIC_KIND_LABEL = {
    "want": "Хотят работать",
    "dont_want": "Не хотят работать",
    "brand_awareness": "Знают компанию",
    "uncertainty": "Не уверены",
}


@dataclass
class SurveyData:
    respondents: pd.DataFrame
    company_raw_long: pd.DataFrame
    company_mapped_long: pd.DataFrame
    factors_long: pd.DataFrame
    rankings: pd.DataFrame
    companies: pd.DataFrame


def normalize_company_name(name: Any) -> str:
    """Нормализует название компании для использования как ключ"""
    s = str(name or "").lower().strip()
    for ch in ["&", "/", "-", "(", ")", ".", ",", "«", "»", "'", "\""]:
        s = s.replace(ch, " ")
    return " ".join(s.split())


def _clean_company_name(name: Any) -> str:
    """Убирает технический префикс ранга вида '83. Company'"""
    s = str(name or "").strip()
    s = re.sub(r"^\d+\.\s*", "", s)
    return s


def _float_or_none(v: Any) -> float | None:
    """Преобразует значение в float или None"""
    try:
        if v in (None, "", "-"):
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _int_or_none(v: Any) -> int | None:
    """Преобразует значение в int или None"""
    f = _float_or_none(v)
    if f is None:
        return None
    return int(f)


def _find_col(df: pd.DataFrame, target: str) -> Any:
    """Находит колонку по точному названию"""
    for c in df.columns:
        if str(c).strip() == target:
            return c
    return None


def _parse_companies_meta(path: Path) -> pd.DataFrame:
    """Парсит метаданные компаний из Excel"""
    wb = load_workbook(path, data_only=True)
    ws = wb["Ответы на форму"]
    rows = []
    for col in range(17, 110):
        company = ws.cell(3, col).value
        sector = ws.cell(2, col).value
        if company in (None, ""):
            continue
        rows.append(
            {
                "company": str(company).strip(),
                "sector": str(sector).strip() if sector not in (None, "") else "Unknown",
                "company_key": normalize_company_name(company),
            }
        )
    companies = pd.DataFrame(rows).drop_duplicates(subset=["company_key"])
    return companies


def _parse_rankings(path: Path) -> pd.DataFrame:
    """Парсит рейтинги компаний из Excel"""
    wb = load_workbook(path, data_only=True)
    ws = wb["Свод (ТОП ВСЕ Компании)"]

    section_starts: list[int] = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().startswith("Сортировка по"):
            section_starts.append(r)

    rows: list[dict[str, Any]] = []
    for i, start in enumerate(section_starts):
        title = str(ws.cell(start, 1).value).strip()
        metric_kind = METRIC_KIND_MAP.get(title)
        if metric_kind is None:
            continue

        data_start = start + 2
        data_end = (section_starts[i + 1] - 1) if i + 1 < len(section_starts) else ws.max_row
        for r in range(data_start, data_end + 1):
            rank_sort = _int_or_none(ws.cell(r, 1).value)
            company = ws.cell(r, 4).value
            sector = ws.cell(r, 2).value
            if rank_sort is None or company in (None, ""):
                continue

            rec = {
                "metric_kind": metric_kind,
                "metric_label": METRIC_KIND_LABEL[metric_kind],
                "company": _clean_company_name(company),
                "company_key": normalize_company_name(_clean_company_name(company)),
                "sector": str(sector).strip() if sector not in (None, "") else "Unknown",
                "rank_sort_2026": rank_sort,
                "company_name_2025": ws.cell(r, 24).value,
                "company_name_2025_key": normalize_company_name(_clean_company_name(ws.cell(r, 24).value)),
                "pct_2025": _float_or_none(ws.cell(r, 26).value),
                "rank_2025": _int_or_none(ws.cell(r, 27).value),
                "pct_2026": _float_or_none(ws.cell(r, 29).value),
                "rank_2026": _int_or_none(ws.cell(r, 30).value),
                "change_pp": _float_or_none(ws.cell(r, 32).value),
                "change_pct_rel": _float_or_none(ws.cell(r, 33).value),
                "change_rank": _int_or_none(ws.cell(r, 34).value),
                "rating_2025_ref": _int_or_none(ws.cell(r, 37).value),
                "change_text_raw": ws.cell(r, 38).value,
            }
            rows.append(rec)
    return pd.DataFrame(rows)


def _parse_rankings_2025_lookup(path: Path) -> pd.DataFrame:
    """Парсит корректные значения 2025 по метрикам из листа 2025."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Свод (ТОП ВСЕ Компании) 2025"]

    section_starts: list[int] = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().startswith("Сортировка по"):
            section_starts.append(r)

    pct_col_by_metric = {
        "want": 15,
        "dont_want": 17,
        "uncertainty": 21,
        "brand_awareness": 23,
    }

    rows: list[dict[str, Any]] = []
    for i, start in enumerate(section_starts):
        title = str(ws.cell(start, 1).value).strip()
        metric_kind = METRIC_KIND_MAP.get(title)
        if metric_kind is None:
            continue

        data_start = start + 2
        data_end = (section_starts[i + 1] - 1) if i + 1 < len(section_starts) else ws.max_row
        pct_col = pct_col_by_metric.get(metric_kind)
        if pct_col is None:
            continue

        for r in range(data_start, data_end + 1):
            rank_2025 = _int_or_none(ws.cell(r, 1).value)
            company = ws.cell(r, 4).value
            if rank_2025 is None or company in (None, ""):
                continue

            company_clean = _clean_company_name(company)
            rows.append(
                {
                    "metric_kind": metric_kind,
                    "company_key": normalize_company_name(company_clean),
                    "company_2025_key": normalize_company_name(company_clean),
                    "company_2025_sheet": company_clean,
                    "pct_2025_fix": _float_or_none(ws.cell(r, pct_col).value),
                    "rank_2025_fix": rank_2025,
                }
            )

    return pd.DataFrame(rows)


def load_survey_data(path: str | Path) -> SurveyData:
    """Загружает все данные опроса из Excel файла"""
    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    # Парсим компании и рейтинги
    companies_meta = _parse_companies_meta(file_path)
    rankings = _parse_rankings(file_path)
    rankings_2025_lookup = _parse_rankings_2025_lookup(file_path)

    # Загружаем ответы
    responses = pd.read_excel(file_path, sheet_name="Ответы на форму", header=2)
    responses = responses.dropna(axis=1, how="all").copy()
    responses = responses[responses.notna().any(axis=1)].copy()

    # Конвертируем числовые колонки
    for c in ["Номер ответа", "Оцените наш опросник"]:
        if c in responses.columns:
            responses[c] = pd.to_numeric(responses[c], errors="coerce")

    if "Номер ответа" in responses.columns:
        responses["Номер ответа"] = responses["Номер ответа"].astype("Int64")

    # Подготавливаем данные респондентов
    respondents = responses.copy()
    respondents["respondent_id"] = respondents["Номер ответа"].astype("Int64")

    # Преобразуем сырые ответы в длинный формат
    raw_company_cols = list(responses.columns[16:109])
    company_raw_long = responses[["Номер ответа"] + raw_company_cols].melt(
        id_vars=["Номер ответа"], value_vars=raw_company_cols, var_name="company", value_name="response_raw"
    )
    company_raw_long = company_raw_long.rename(columns={"Номер ответа": "respondent_id"})
    company_raw_long["respondent_id"] = company_raw_long["respondent_id"].astype("Int64")
    company_raw_long["company"] = company_raw_long["company"].astype(str).str.strip()
    company_raw_long["company_key"] = company_raw_long["company"].map(normalize_company_name)
    company_raw_long["response_mapped"] = company_raw_long["response_raw"].map(RAW_RESPONSE_TO_MAPPED)

    # Добавляем информацию о секторе
    company_raw_long = company_raw_long.merge(
        companies_meta[["company_key", "sector"]], on="company_key", how="left"
    )
    company_raw_long["sector"] = company_raw_long["sector"].fillna("Unknown")

    # Преобразуем маппированные ответы в длинный формат
    mapped_company_cols = [c for c in responses.columns[111:205] if str(c) != "открытый вопрос"]
    company_mapped_long = responses[["Номер ответа"] + mapped_company_cols].melt(
        id_vars=["Номер ответа"], value_vars=mapped_company_cols, var_name="company_mapped_col", value_name="response_mapped_sheet"
    )
    company_mapped_long = company_mapped_long.rename(columns={"Номер ответа": "respondent_id"})
    company_mapped_long["respondent_id"] = company_mapped_long["respondent_id"].astype("Int64")
    company_mapped_long["company"] = (
        company_mapped_long["company_mapped_col"].astype(str).str.replace(".1", "", regex=False).str.strip()
    )
    company_mapped_long["company_key"] = company_mapped_long["company"].map(normalize_company_name)

    # Парсим факторы
    factor_defs = [
        ("top_factors", ["1", "2", "3"]),
        ("offer_reject_factors", ["1.1", "2.1", "3.1"]),
        ("sources", ["1.2", "2.2", "3.2"]),
        ("job_change_reasons", ["1.3", "2.3", "3.3"]),
    ]
    factor_rows: list[dict[str, Any]] = []
    for group_name, ranks in factor_defs:
        for pos, rank in enumerate(ranks, start=1):
            col = _find_col(responses, rank)
            if col is None:
                continue
            s = responses[["Номер ответа", col]].rename(columns={"Номер ответа": "respondent_id", col: "factor"})
            s["factor_group"] = group_name
            s["rank_position"] = pos
            factor_rows.append(s)
    
    factors_long = pd.concat(factor_rows, ignore_index=True) if factor_rows else pd.DataFrame()
    if not factors_long.empty:
        factors_long["respondent_id"] = factors_long["respondent_id"].astype("Int64")
        factors_long = factors_long[factors_long["factor"].notna()].copy()
        factors_long["factor"] = factors_long["factor"].astype(str).str.strip()
        factors_long = factors_long[factors_long["factor"] != ""].copy()

    # Добавляем информацию о секторе в рейтинги
    if not rankings.empty:
        if not rankings_2025_lookup.empty:
            rankings = rankings.merge(
                rankings_2025_lookup[
                    ["metric_kind", "company_2025_key", "pct_2025_fix", "rank_2025_fix"]
                ],
                left_on=["metric_kind", "company_name_2025_key"],
                right_on=["metric_kind", "company_2025_key"],
                how="left",
            )
            rankings["pct_2025"] = rankings["pct_2025_fix"].combine_first(rankings["pct_2025"])
            rankings["rank_2025"] = rankings["rank_2025_fix"].combine_first(rankings["rank_2025"])
            rankings = rankings.drop(columns=["pct_2025_fix", "rank_2025_fix", "company_2025_key"])

        # Пересчитываем производные показатели на основе скорректированных значений 2025.
        mask_pct = rankings["pct_2025"].notna() & rankings["pct_2026"].notna()
        rankings.loc[mask_pct, "change_pp"] = rankings.loc[mask_pct, "pct_2026"] - rankings.loc[mask_pct, "pct_2025"]

        mask_rel = mask_pct & rankings["pct_2025"].ne(0)
        rankings.loc[mask_rel, "change_pct_rel"] = rankings.loc[mask_rel, "change_pp"] / rankings.loc[mask_rel, "pct_2025"]
        rankings.loc[mask_pct & rankings["pct_2025"].eq(0), "change_pct_rel"] = pd.NA

        mask_rank = rankings["rank_2025"].notna() & rankings["rank_2026"].notna()
        rankings.loc[mask_rank, "change_rank"] = rankings.loc[mask_rank, "rank_2025"] - rankings.loc[mask_rank, "rank_2026"]

        rankings = rankings.merge(
            companies_meta[["company_key", "sector"]].rename(columns={"sector": "sector_meta"}),
            on="company_key",
            how="left",
        )
        rankings["sector"] = rankings["sector"].fillna(rankings["sector_meta"]).fillna("Unknown")
        rankings = rankings.drop(columns=["sector_meta"])

    # Создаём справочник компаний
    companies = (
        rankings[["company", "company_key", "sector"]]
        .drop_duplicates(subset=["company_key"])
        .sort_values(["sector", "company"])
        .reset_index(drop=True)
    )

    return SurveyData(
        respondents=respondents,
        company_raw_long=company_raw_long,
        company_mapped_long=company_mapped_long,
        factors_long=factors_long,
        rankings=rankings,
        companies=companies,
    )


def apply_respondent_filters(
    respondents: pd.DataFrame,
    selected_gender: list[str] | None = None,
    selected_age: list[str] | None = None,
    selected_experience: list[str] | None = None,
    selected_specialization: list[str] | None = None,
) -> pd.DataFrame:
    """Применяет фильтры к данным респондентов"""
    df = respondents.copy()
    if selected_gender:
        df = df[df["Гендер"].isin(selected_gender)]
    if selected_age:
        df = df[df["Возраст"].isin(selected_age)]
    if selected_experience:
        df = df[df["Опыт работы в ИТ / Digital"].isin(selected_experience)]
    if selected_specialization:
        df = df[df["К какой специализации Вы себя относите?"].isin(selected_specialization)]
    return df
